# -*- coding: utf-8 -*-
"""
Exportación de una ejecución de OF (finalizada o cancelada) a un Excel
acumulativo. El fichero se guarda en EXPORT_FOLDER, que en producción debe
ser una carpeta sincronizada con la nube (OneDrive/Google Drive/SharePoint)
para que quede disponible para fabricacion@quiadsa.com.

Si se dispone de integración directa (p.ej. Microsoft Graph API para subir
a OneDrive/SharePoint), se puede añadir esa llamada al final de
`exportar_ejecucion()` (ver TODO al final del fichero).
"""

import os
import datetime
import openpyxl
from openpyxl.utils import get_column_letter

import config

COLUMNAS = [
    "Fecha/Hora exportación",
    "Estado",
    "Código OF",
    "Máquina",
    "Operarios iniciales",
    "Fecha/Hora inicio OF",
    "Fecha/Hora fin OF",
    "Tiempo total (min)",
    "Peso tara (kg)",
    "Peso total (kg)",
    "Peso neto (kg)",
    "Incidencia peso neto",
    "Sección Nº",
    "Instrucción sección",
    "Operarios sección",
    "Tiempo previsto sección (min)",
    "Tiempo real sección (min)",
    "Incidencia sección",
    "Pausa inicio",
    "Pausa fin",
    "Pausa (min)",
]


def _asegurar_carpeta():
    os.makedirs(config.EXPORT_FOLDER, exist_ok=True)
    return os.path.join(config.EXPORT_FOLDER, config.EXCEL_FILENAME)


def _abrir_o_crear_libro(ruta_excel):
    if os.path.exists(ruta_excel):
        wb = openpyxl.load_workbook(ruta_excel)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registro"
        ws.append(COLUMNAS)
        for i, _ in enumerate(COLUMNAS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 22
    return wb, ws


def _pausas_de_seccion(pausas: list, seccion_id) -> tuple:
    """Filtra las pausas que ocurrieron durante una sección concreta y
    devuelve tres cadenas (inicio, fin, minutos), una por columna. Si hay
    varias pausas en la misma sección, se separan con ' ; ' dentro de cada
    columna, manteniendo el mismo orden en las tres."""
    de_la_seccion = [p for p in pausas if p.get("ejecucion_seccion_id") == seccion_id]
    if not de_la_seccion:
        return "", "", ""
    inicios = [p.get("fecha_inicio") or "" for p in de_la_seccion]
    fines = [p.get("fecha_fin") or "(sin finalizar)" for p in de_la_seccion]
    minutos = [str(p.get("tiempo_total_min")) if p.get("tiempo_total_min") is not None
               else "-" for p in de_la_seccion]
    return " ; ".join(inicios), " ; ".join(fines), " ; ".join(minutos)


def exportar_ejecucion(ejecucion: dict, secciones: list, pausas: list = None):
    """
    Añade al Excel una fila por cada sección de la ejecución indicada.
    `ejecucion` es el dict devuelto por LocalRepository.obtener_ejecucion().
    `secciones` es la lista de dicts devuelta por obtener_secciones()
    (incluye la clave "operarios" con los operarios de cada sección).
    `pausas` es la lista de dicts devuelta por obtener_pausas(); cada pausa
    incluye "ejecucion_seccion_id" para poder mostrarla en la fila de la
    sección durante la que ocurrió.
    """
    ruta_excel = _asegurar_carpeta()
    wb, ws = _abrir_o_crear_libro(ruta_excel)

    fecha_exportacion = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    peso_tara = ejecucion.get("peso_tara")
    peso_llena = ejecucion.get("peso_llena")
    peso_neto = None
    if peso_tara is not None and peso_llena is not None:
        peso_neto = round(peso_llena - peso_tara, 2)

    pausas = pausas or []
    incidencia_peso = ejecucion.get("peso_incidencia_motivo") or ""

    fila_comun = [
        fecha_exportacion,
        ejecucion["estado"],
        ejecucion["codigo_of"],
        ejecucion["maquina_id"],
        ejecucion["operarios"],
        ejecucion["fecha_inicio"],
        ejecucion["fecha_fin"],
        ejecucion["tiempo_total_min"],
        peso_tara,
        peso_llena,
        peso_neto,
        incidencia_peso,
    ]

    if secciones:
        for sec in secciones:
            pausa_inicio, pausa_fin, pausa_min = _pausas_de_seccion(pausas, sec["id"])
            ws.append(fila_comun + [
                sec["numero_seccion"],
                sec["texto"],
                ", ".join(sec.get("operarios", [])),
                sec["tiempo_previsto_min"],
                sec["tiempo_real_min"],
                sec["incidencia_motivo"] or "",
                pausa_inicio,
                pausa_fin,
                pausa_min,
            ])
    else:
        # Cancelada sin ninguna sección iniciada.
        ws.append(fila_comun + ["", "", "", "", "", "", "", "", ""])

    wb.save(ruta_excel)
    return ruta_excel

    # -----------------------------------------------------------------
    # TODO (opcional): Subida directa a la nube vía API en lugar de una
    # carpeta local sincronizada. Ejemplo con Microsoft Graph (OneDrive):
    #
    # import requests
    # token = obtener_token_graph()  # OAuth2 con la cuenta fabricacion@quiadsa.com
    # with open(ruta_excel, "rb") as f:
    #     requests.put(
    #         f"https://graph.microsoft.com/v1.0/me/drive/root:/{config.EXCEL_FILENAME}:/content",
    #         headers={"Authorization": f"Bearer {token}"},
    #         data=f,
    #     )
    # -----------------------------------------------------------------
